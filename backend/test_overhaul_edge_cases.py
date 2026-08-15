import pytest
from backend.price_parser import extract_current_a_from_name, parse_price_list_raw, clean_key
from backend.pdf_parser import is_trash_item, text_fallback_scheme_parser

def test_extract_current_a_from_name():
    # 1. Ignore series frame sizes (250 is frame size, 125 is current_a)
    assert extract_current_a_from_name("NM8N-250H 3P 125А") == 125
    assert extract_current_a_from_name("NM8N-250H 3P 125A") == 125

    # 2. Ignore breaking capacities (36кА is breaking capacity, 63 is current_a)
    assert extract_current_a_from_name("NM8N-250H 3P 63А 36кА") == 63
    assert extract_current_a_from_name("NM8N-250H 3P 63А 100kA") == 63

    # 3. Prefer rating value closer to poles indicator
    # "3P" is at index 11. "125A" is closer to "3P" than any other number.
    assert extract_current_a_from_name("NM8N-250H 3P 125A") == 125
    assert extract_current_a_from_name("125A 3P") == 125

def test_is_trash_item():
    assert is_trash_item("мм") is True
    assert is_trash_item("адрес") is True
    assert is_trash_item("QF1") is True  # standard QF designation without nominal rating
    assert is_trash_item("QF1", "C16") is False  # has nominal rating
    assert is_trash_item("Авт. выкл. 3P 125А") is False  # long enough and has nominal rating
    assert is_trash_item("QF", "") is True  # short and no rating

def test_parse_price_list_raw_overlapping_bug():
    # Create mock excel data to test raw price parser
    import openpyxl
    import io

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Артикул", "Наименование", "Тариф с НДС, руб"])
    ws.append(["12345", "Авт. выкл. 3P 125A", "1000.50"])
    ws.append(["67890", "Авт. выкл. 3P 63A", "67890"])  # price equals article -> should trigger warning and price=0.0

    out = io.BytesIO()
    wb.save(out)

    raw_list = parse_price_list_raw(out.getvalue())
    # Exclude header row from counting
    content_rows = [r for r in raw_list if r["Артикул"] != "Артикул"]
    assert len(content_rows) == 2
    assert content_rows[0]["price"] == 1000.50
    assert content_rows[1]["price"] == 0.0  # Safe recovery from price overlap bug!

def test_detect_columns_multi_vat_prioritization():
    from backend.price_parser import detect_columns

    # Multi-column table with both 'Тариф без НДС, руб' and 'Тариф с НДС, руб'
    rows = [
        ("Код", "Наименование номенклатуры", "Тариф с НДС, руб", "Тариф без НДС, руб", "Ед. изм."),
        ("268974", "Авт. выкл. NM8N-400H TM 3P 400A 100kA", "58670.85", "48090.86", "шт"),
        ("268975", "Авт. выкл. NM8N-400H TM 3P 315A 100kA", "55000.00", "45000.00", "шт")
    ]
    art_idx, name_idx, price_idx, is_kopecks = detect_columns(rows)
    assert art_idx == 0
    assert name_idx == 1
    assert price_idx == 2  # Must pick column index 2 ('Тариф с НДС, руб') over column index 3 ('Тариф без НДС, руб')
    assert is_kopecks is False

def test_detect_columns_single_tariff_without_vat():
    from backend.price_parser import detect_columns

    # Single price column table where header is 'Тариф без НДС, руб' (and numeric cols like weight exist)
    rows = [
        ("Артикул", "Наименование", "Тариф без НДС, руб", "Масса, кг", "Объем, куб.м"),
        ("268974", "Авт. выкл. NM8N-400H TM 3P 400A", "48090.86", "5.735", "0.011"),
        ("268975", "Авт. выкл. NM8N-400H TM 3P 315A", "45000.00", "5.240", "0.011")
    ]
    art_idx, name_idx, price_idx, is_kopecks = detect_columns(rows)
    assert art_idx == 0
    assert name_idx == 1
    assert price_idx == 2  # Must pick 'Тариф без НДС, руб' as the price column
    assert is_kopecks is False

def test_currency_rub_header_detection():
    from backend.price_parser import detect_columns

    # Table where the price header is simply 'Руб.'
    rows = [
        ("Код", "Наименование", "Руб.", "Ед. изм."),
        ("268974", "Авт. выкл. NM8N-400H 3P 400A", "58670.85", "шт")
    ]
    art_idx, name_idx, price_idx, is_kopecks = detect_columns(rows)
    assert art_idx == 0
    assert name_idx == 1
    assert price_idx == 2  # Must identify 'Руб.' as the price column
    assert is_kopecks is False

def test_chint_image_price_list_structure():
    from backend.price_parser import detect_columns, parse_price_list
    import openpyxl, io

    # Real CHINT price list format matching the user image structure with metadata columns
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Код", "Наименование номенклатуры", "Тариф с НДС, руб", "Тариф без НДС, руб", "Ед. изм.", "Тип", "Склад", "Вес, кг"])
    for _ in range(50):
        ws.append(["268974", "Авт. выкл. NM8N-400H TM 3P 400A 100kA с рег. термомаг. расцепителем (R)", "58 670,85", "48 090,86", "шт", "Промышленная", "Да", "5,735"])

    out = io.BytesIO()
    wb.save(out)

    rows = list(openpyxl.load_workbook(out, data_only=True).active.iter_rows(values_only=True))
    art_idx, name_idx, price_idx, is_kopecks = detect_columns(rows)

    assert art_idx == 0    # Must be "Код" column
    assert name_idx == 1   # Must be "Наименование" column
    assert price_idx == 2  # Must be "Тариф с НДС, руб" column

    out.seek(0)
    pm = parse_price_list(out.getvalue())
    assert "268974" in pm
    assert pm["268974"] == 58670.85

def test_detect_columns_robustness():
    from backend.price_parser import detect_columns

    # Custom Excel table #1: Russian alternate headers
    rows_1 = [
        ("Код товара", "Название позиции", "Цена"),
        ("12345", "Автоматический выключатель CHINT NB2 3P 16A", "150.00"),
        ("67890", "Автоматический выключатель CHINT NB2 3P 32A", "250.00")
    ]
    art_idx, name_idx, price_idx, is_kopecks = detect_columns(rows_1)
    assert art_idx == 0
    assert name_idx == 1
    assert price_idx == 2
    assert is_kopecks is False

    # Custom Excel table #2: Scrambled columns & English headers
    rows_2 = [
        ("Price", "Code", "Item Description"),
        ("150.00", "12345", "Автоматический выключатель CHINT NB2 3P 16A"),
        ("250.00", "67890", "Автоматический выключатель CHINT NB2 3P 32A")
    ]
    art_idx, name_idx, price_idx, is_kopecks = detect_columns(rows_2)
    assert art_idx == 1
    assert name_idx == 2
    assert price_idx == 0
    assert is_kopecks is False

    # Custom Excel table #3: Missing headers completely (uses numeric fallback & content analysis)
    # The first row is raw values, not headers.
    rows_3 = [
        ("99999", "Автоматический выключатель CHINT DZ158 3P 100A", "1200.50"),
        ("88888", "Автоматический выключатель CHINT DZ158 3P 125A", "1500.00")
    ]
    art_idx, name_idx, price_idx, is_kopecks = detect_columns(rows_3)
    assert art_idx == 0
    assert name_idx == 1
    assert price_idx == 2
    assert is_kopecks is False

def test_parse_robust_float():
    from backend.price_parser import parse_robust_float
    assert parse_robust_float("1.500,50") == 1500.5
    assert parse_robust_float("150,50 руб.") == 150.5
    assert parse_robust_float("1500.00") == 1500.0
    assert parse_robust_float("1 500") == 1500.0
    assert parse_robust_float(None) == 0.0
    assert parse_robust_float("") == 0.0

def test_kopecks_auto_division():
    from backend.price_parser import detect_columns, parse_price_list_raw
    import openpyxl
    import io

    # Create mock excel with prices in kopecks (e.g. 15050 kopecks represents 150.50 rubles)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Артикул", "Наименование номенклатуры", "Цена, в копейках"])
    ws.append(["11111", "Автоматический выключатель 3P 16A", "15050"])
    ws.append(["22222", "Автоматический выключатель 1P 10A", "7500"])

    out = io.BytesIO()
    wb.save(out)

    # 1. Test column detection says is_kopecks is True
    rows = list(openpyxl.load_workbook(out, data_only=True).active.iter_rows(values_only=True))
    art_idx, name_idx, price_idx, is_kopecks = detect_columns(rows)
    assert is_kopecks is True

    # 2. Test raw parse auto-divides prices by 100.0
    out.seek(0)
    raw_list = parse_price_list_raw(out.getvalue())
    content_rows = [r for r in raw_list if r["Артикул"] != "Артикул"]
    assert content_rows[0]["price"] == 150.50
    assert content_rows[1]["price"] == 75.00


def test_text_fallback_scheme_parser_parallel_sequences():
    # Simulate parallel sequence lists from diagrams
    input_text = """
    125А 125А 125А 125А 125А 63А 63А 16А 16А 125А 125А 630А
    3P 3P 3P 3P 3P 1P 1P 1P 1P 3P 3P 3P
    ввод 504А
    """
    items = text_fallback_scheme_parser(input_text)

    # 504A should be ignored because 630A is present in the text!
    assert not any(it["current_a"] == "504" for it in items)

    # 630A 3P should be mapped correctly (qty=1)
    item_630 = next(it for it in items if it["current_a"] == "630")
    assert item_630["poles"] == "3P"
    assert item_630["qty"] == 1

    # Check 125A mapping (qty=7)
    item_125 = next(it for it in items if it["current_a"] == "125")
    assert item_125["poles"] == "3P"
    assert item_125["qty"] == 7

    # Check 63A mapping (qty=2)
    item_63 = next(it for it in items if it["current_a"] == "63")
    assert item_63["poles"] == "1P"
    assert item_63["qty"] == 2

    # Check 16A mapping (qty=2)
    item_16 = next(it for it in items if it["current_a"] == "16")
    assert item_16["poles"] == "1P"
    assert item_16["qty"] == 2
