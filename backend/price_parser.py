import openpyxl
from typing import Dict, Any, List, Tuple, Optional
import io
import re
import logging

logger = logging.getLogger("price_parser")

def clean_key(val: Any) -> str:
    """Cleans up a code/article/name to allow flexible matching (removes hyphens, spaces, slashes, uppercase)."""
    if val is None:
        return ""
    # Convert to uppercase string, strip spaces, remove hyphens and slashes
    s = str(val).strip().upper()
    return re.sub(r'[^A-Z0-9А-Я]', '', s)

def parse_robust_float(val: Any) -> float:
    """
    Intelligently and robustly parses pricing values into floats, handling
    varying formats of thousands and decimal separators (Russian/English/German).
    """
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)

    s = str(val).strip()
    if not s:
        return 0.0

    # Strip spaces and currency symbols
    s = re.sub(r'[\s\xa0\u200b\u202f\tруб\$€₽]+', '', s, flags=re.IGNORECASE)
    # Strip trailing dot from abbreviations like 'руб.'
    s = s.rstrip('.')
    if not s:
        return 0.0

    if '.' in s and ',' in s:
        dot_idx = s.find('.')
        comma_idx = s.find(',')
        if dot_idx < comma_idx:
            # e.g. 1.500,50 -> dot is thousands, comma is decimal
            s = s.replace('.', '').replace(',', '.')
        else:
            # e.g. 1,500.50 -> comma is thousands, dot is decimal
            s = s.replace(',', '')
    elif ',' in s:
        # In Russian spreadsheets, comma is almost always decimal separator
        s = s.replace(',', '.')
    elif '.' in s:
        if s.count('.') > 1:
            # Multiple dots are thousands separators, e.g. 1.500.000 -> 1500000
            s = s.replace('.', '')
        # Note: Avoid stripping a single dot for 3 decimals (e.g. 150.000 or 150.250)
        # to prevent corrupting numbers formatted with 3 decimal places.

    s_clean = ''
    has_dot = False
    for c in s:
        if c.isdigit():
            s_clean += c
        elif c == '-' and not s_clean:
            s_clean += c
        elif c == '.' and not has_dot:
            s_clean += c
            has_dot = True
    try:
        return float(s_clean) if s_clean else 0.0
    except ValueError:
        return 0.0

def extract_current_a_from_name(name_str: str) -> Optional[int]:
    """
    Intelligently extracts the nominal current rating (current_a) from name_str.
    - Ignores breaking capacities (e.g. 100кА, 36кА).
    - Ignores series frame sizes (e.g. NM8N-250H).
    - Prefers rating value closer to the poles indicator (e.g. 3P 63А -> 63).
    """
    if not name_str:
        return None

    # 1. Clean breaking capacity (кА / kA)
    name_clean = re.sub(r'\b\d+\s*(?:кА|kA|ка|ka)\b', '', name_str, flags=re.IGNORECASE)

    # 1. Clean product series frame sizes (e.g. NM8N-250, NXM-160S, NXB-63, DZ158-125)
    name_clean = re.sub(r'\b(?:NM8[N,S]|NXM|NM1|NXB|NB[1,2,8]|NC[1,2,7,8]|NVF[2,5,7]|NZ7|NKB1|DZ158|NR[8,E]|NL1|NH4|ND2)-\d+[A-Z0-9/-]*\b', '', name_clean, flags=re.IGNORECASE)

    # Find all poles positions (1P/2P/3P/4P)
    poles_indices = []
    for m in re.finditer(r'\b([1-4])\s*(?:P|П|полюс|п|p)\b', name_clean, re.IGNORECASE):
        poles_indices.append(m.start())

    # Find all amperage rating candidates
    amp_candidates = []
    for m in re.finditer(r'\b(\d+)\s*(?:А|A)\b', name_clean, re.IGNORECASE):
        try:
            val = int(m.group(1))
            amp_candidates.append((val, m.start()))
        except ValueError:
            pass

    if not amp_candidates:
        return None

    if len(amp_candidates) == 1 or not poles_indices:
        return amp_candidates[0][0]

    # Choose candidate closest to any poles indicator
    best_val = amp_candidates[0][0]
    min_dist = float('inf')
    for val, start_idx in amp_candidates:
        for p_idx in poles_indices:
            dist = abs(start_idx - p_idx)
            if dist < min_dist:
                min_dist = dist
                best_val = val

    return best_val

def detect_columns(rows: List[Tuple[Any, ...]]) -> Tuple[int, int, int, bool]:
    """
    Intelligently and robustly detects column indices for:
    - Article (Артикул)
    - Name (Наименование)
    - Price (Тариф с НДС, руб / Цена)
    And detects if the price column lists values in kopecks (is_kopecks).
    Returns a tuple of (col_article_idx, col_name_idx, col_price_idx, is_kopecks).
    """
    if not rows:
        return 0, 1, 2, False

    num_cols = len(rows[0])
    if num_cols == 0:
        return 0, 1, 2, False

    # 1. Exact Header Row Search by finding the row containing BOTH "артикул" and "наименование"
    header_r_idx = -1
    for r_idx in range(min(15, len(rows))):
        row = rows[r_idx]
        row_strs = [str(cell).strip().lower() for cell in row if cell is not None]
        has_art = any('артикул' in s for s in row_strs)
        has_name = any('наименование' in s for s in row_strs)
        if has_art and has_name:
            header_r_idx = r_idx
            logger.info(f"[Detect Columns] Found exact header row at row index {header_r_idx}: {row}")
            break

    if header_r_idx == -1:
        # Fallback search if exact pair is not in one row
        for r_idx in range(min(15, len(rows))):
            row = rows[r_idx]
            row_strs = [str(cell).strip().lower() for cell in row if cell is not None]
            has_art = any(any(kw in s for kw in ['артикул', 'код', 'sku', 'article']) for s in row_strs)
            has_price = any(any(kw in s for kw in ['цена', 'тариф', 'стоимость', 'price']) for s in row_strs)
            has_name = any(any(kw in s for kw in ['наимен', 'назван', 'товар', 'описан', 'name', 'item']) for s in row_strs)
            if (has_art and has_price) or (has_name and has_price) or (has_art and has_name):
                header_r_idx = r_idx
                break

    article_scores = [0] * num_cols
    name_scores = [0] * num_cols
    price_scores = [0] * num_cols

    # Multi-row combined header scan per column (prioritizing header_r_idx row)
    max_header_rows = min(10, len(rows))
    for c_idx in range(num_cols):
        cell_header_exact = ""
        if header_r_idx != -1 and c_idx < len(rows[header_r_idx]) and rows[header_r_idx][c_idx] is not None:
            cell_header_exact = str(rows[header_r_idx][c_idx]).strip().lower()

        col_text = " ".join(
            str(rows[r_idx][c_idx]).strip().lower()
            for r_idx in range(max_header_rows)
            if c_idx < len(rows[r_idx]) and rows[r_idx][c_idx] is not None
        )
        if not col_text and not cell_header_exact:
            continue

        text_to_check = cell_header_exact if cell_header_exact else col_text

        # Blacklisted columns: NEVER select as article or price
        if any(kw in text_to_check for kw in ['коллекция', 'складской', 'склад', 'группа', 'серия', 'вес', 'объём', 'объем', 'категория', 'тип', 'ед.', 'ед.изм', 'ед. изм', 'аналог']):
            article_scores[c_idx] -= 10000
            name_scores[c_idx] -= 10000
            price_scores[c_idx] -= 10000

        # Exact Article matching: contains "артикул" and NOT "аналог"
        if 'артикул' in text_to_check and 'аналог' not in text_to_check:
            article_scores[c_idx] += 10000
        elif any(kw in text_to_check for kw in ['код товара', 'код номенклатуры', 'код']) and not any(kw in text_to_check for kw in ['руб', 'цена', 'тариф', 'аналог']):
            article_scores[c_idx] += 5000

        # Exact Name matching: contains "наименование"
        if 'наименование' in text_to_check and not any(kw in text_to_check for kw in ['код', 'артикул', 'руб', 'цена', 'тариф']):
            name_scores[c_idx] += 10000
        elif any(kw in text_to_check for kw in ['номенклатура', 'название', 'описание']) and not any(kw in text_to_check for kw in ['код', 'артикул', 'руб']):
            name_scores[c_idx] += 5000

        # Strict Price matching hierarchy:
        # 1) "тариф с ндс"
        # 2) "цена с ндс"
        # 3) "тариф" + "ндс"
        if 'тариф с ндс' in text_to_check:
            price_scores[c_idx] += 10000
        elif 'цена с ндс' in text_to_check:
            price_scores[c_idx] += 8000
        elif 'тариф' in text_to_check and 'ндс' in text_to_check and 'без' not in text_to_check:
            price_scores[c_idx] += 7000
        elif any(kw in text_to_check for kw in ['цена', 'тариф', 'стоимость', 'руб', '₽']) and 'без ндс' not in text_to_check:
            price_scores[c_idx] += 3000
        elif 'тариф без ндс' in text_to_check or 'цена без ндс' in text_to_check:
            price_scores[c_idx] += 1500

    # 2. Content analysis on remaining rows (capped to max 50 points total to avoid overriding headers)
    start_content_row = header_r_idx + 1 if header_r_idx != -1 else 0
    raw_content_art = [0] * num_cols
    raw_content_name = [0] * num_cols
    raw_content_price = [0] * num_cols

    for r_idx in range(start_content_row, min(start_content_row + 30, len(rows))):
        row = rows[r_idx]
        for c_idx in range(min(num_cols, len(row))):
            val = row[c_idx]
            if val is None:
                continue
            val_str = str(val).strip()
            if not val_str:
                continue

            is_num = False
            clean_str = re.sub(r'[\s\xa0\u200b\u202f\tруб\$€₽]+', '', val_str).replace(',', '.')
            if re.match(r'^\d+(\.\d+)?$', clean_str):
                is_num = True

            if is_num:
                if '.' in clean_str:
                    raw_content_price[c_idx] += 5
                else:
                    raw_content_price[c_idx] += 1
                    raw_content_art[c_idx] += 1
            else:
                if len(val_str) > 15 and ' ' in val_str:
                    raw_content_name[c_idx] += 5
                elif 3 <= len(val_str) <= 25:
                    raw_content_art[c_idx] += 3

    for c_idx in range(num_cols):
        article_scores[c_idx] += min(raw_content_art[c_idx], 50)
        name_scores[c_idx] += min(raw_content_name[c_idx], 50)
        price_scores[c_idx] += min(raw_content_price[c_idx], 50)

    # Selection logic
    best_price_idx = 0
    max_price = -1
    for c in range(num_cols):
        if price_scores[c] > max_price:
            max_price = price_scores[c]
            best_price_idx = c

    best_name_idx = 0
    max_name = -1
    for c in range(num_cols):
        if c == best_price_idx:
            continue
        if name_scores[c] > max_name:
            max_name = name_scores[c]
            best_name_idx = c

    best_article_idx = 0
    max_article = -1
    for c in range(num_cols):
        if c == best_price_idx or c == best_name_idx:
            continue
        if article_scores[c] > max_article:
            max_article = article_scores[c]
            best_article_idx = c

    if num_cols == 1:
        return 0, 0, 0, False
    elif num_cols == 2:
        return 0, best_name_idx, best_price_idx, False

    # Analyze if price header indicates kopecks
    is_kopecks = False
    if header_r_idx != -1 and best_price_idx < len(rows[header_r_idx]):
        price_header_cell = rows[header_r_idx][best_price_idx]
        if price_header_cell is not None:
            header_str = str(price_header_cell).strip().lower()
            if any(kw in header_str for kw in ["коп", "копеек", "копейки", "копейках"]) and not "копир" in header_str:
                is_kopecks = True
                logger.info(f"[Price Parser] Kopecks auto-division detected on header: '{header_str}'")

    logger.info(f"[Detect Columns] Selected: Article={best_article_idx}, Name={best_name_idx}, Price={best_price_idx}, is_kopecks={is_kopecks}")
    return best_article_idx, best_name_idx, best_price_idx, is_kopecks


def parse_price_list(file_bytes: bytes, price_map: Optional[Dict[str, float]] = None) -> Dict[str, float]:
    """
    Parses a Price List Excel file (.xlsx) and returns a dictionary (or updates existing dictionary)
    mapping cleaned articles or cleaned names to prices with VAT.
    """
    if price_map is None:
        price_map = {}

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet = wb.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return price_map

    col_article_idx, col_name_idx, col_price_idx, is_kopecks = detect_columns(rows)
    logger.info(f"[Price Parser] col_article_idx={col_article_idx}, col_name_idx={col_name_idx}, col_price_idx={col_price_idx}, is_kopecks={is_kopecks}")

    for row in rows:
        if not any(row):
            continue

        art_val = row[col_article_idx] if col_article_idx < len(row) else None
        name_val = row[col_name_idx] if col_name_idx < len(row) else None
        price_val = row[col_price_idx] if col_price_idx < len(row) else None

        if not art_val and not name_val:
            continue

        # Parse price with robust float parser
        price = parse_robust_float(price_val)
        if is_kopecks:
            price /= 100.0

        price_str = str(price_val).strip() if price_val is not None else ""
        article_str = str(art_val).strip() if art_val is not None else ""

        # Validation rule 1: Verify article is not pure non-numeric category text (e.g. "Промышленная")
        if article_str and not re.search(r'\d', article_str):
            logger.error(f"[Price Validation Error] Article '{article_str}' is pure text without digits. Clearing article.")
            article_str = ""

        # Validation rule 2: Check if price equals article code (price == article)
        if article_str and (price_str == article_str or clean_key(article_str) == clean_key(price_str) or (price > 0 and str(int(price)) == article_str)):
            logger.error(f"[Price Validation Error] Parsed price ({price}) equals article code ({article_str}). Setting price to 0.0.")
            price = 0.0

        if article_str and price > 0.0:
            price_map[clean_key(article_str)] = price
        if name_val and price > 0.0:
            price_map[clean_key(name_val)] = price

    return price_map

def parse_excel_to_unified(file_bytes: bytes) -> List[Dict[str, Any]]:
    """
    Parses price list Excel and extracts positions into a unified format list.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet = wb.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    col_article_idx, col_name_idx, col_price_idx, is_kopecks = detect_columns(rows)
    logger.info(f"[Price Parser Unified] col_article_idx={col_article_idx}, col_name_idx={col_name_idx}, col_price_idx={col_price_idx}, is_kopecks={is_kopecks}")

    unified_items = []

    for row in rows:
        if not any(row):
            continue

        art_val = row[col_article_idx] if col_article_idx < len(row) else None
        name_val = row[col_name_idx] if col_name_idx < len(row) else None
        price_val = row[col_price_idx] if col_price_idx < len(row) else None

        if not name_val:
            continue

        # Skip header row if name_val matches column title
        if str(name_val).strip().lower() in ["наименование", "наименование номенклатуры", "название", "номенклатура"]:
            continue

        # Parse price with robust float parser
        price = parse_robust_float(price_val)
        if is_kopecks:
            price /= 100.0

        price_str = str(price_val).strip() if price_val is not None else ""
        name_str = str(name_val).strip()
        article_str = str(art_val).strip() if art_val is not None else ""

        # Validation rule 1: Verify article is not pure non-numeric category text (e.g. "Промышленная")
        if article_str and not re.search(r'\d', article_str):
            logger.error(f"[Price Validation Error] Article '{article_str}' is pure text without digits. Clearing article.")
            article_str = ""

        # Validation rule 2: Check if price equals article code (price == article)
        if article_str and (price_str == article_str or clean_key(article_str) == clean_key(price_str) or (price > 0 and str(int(price)) == article_str)):
            logger.error(f"[Price Validation Error] Parsed price ({price}) equals article code ({article_str}). Setting price to 0.0.")
            price = 0.0

        # 1. Extract poles
        poles = None
        poles_match = re.search(r'\b([1-4])\s*(?:P|П|полюс|п|p)\b', name_str, re.IGNORECASE)
        if poles_match:
            poles = f"{poles_match.group(1)}P"

        # 2. Extract current_a using the highly intelligent amperage extractor
        current_a = extract_current_a_from_name(name_str)

        # 3. Extract series (expanded pattern covering CHINT & common electrical breaker families)
        series = None
        series_match = re.search(r'\b((?:NM8[N,S]|NXM|NM1|NXB|NB[1,2,8]|NC[1,2,7,8]|NVF[2,5,7]|NZ7|NKB1|DZ158|NR[8,E]|NL1|NH4|ND2)[-A-Z0-9/]*)\b', name_str, re.IGNORECASE)
        if series_match:
            series = series_match.group(1).upper()

        unified_items.append({
            "article": article_str,
            "name": name_str,
            "price": price,
            "current_a": current_a,
            "poles": poles,
            "series": series
        })

    return unified_items

def build_and_save_index(file_bytes: bytes) -> Dict[str, Any]:
    """
    Builds structured index {poles}_{current_a} -> list of unified positions.
    """
    unified_items = parse_excel_to_unified(file_bytes)
    index_map = {}
    for item in unified_items:
        poles = item.get("poles")
        current_a = item.get("current_a")
        if poles and current_a:
            key = f"{poles}_{current_a}"
            if key not in index_map:
                index_map[key] = []
            index_map[key].append(item)
    return index_map

def parse_price_list_raw(file_bytes: bytes) -> List[Dict[str, Any]]:
    """
    Parses a Price List Excel file and returns a list of dictionaries with original keys:
    'Артикул', 'Наименование', 'Тариф с НДС, руб', and 'price'.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet = wb.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    col_article_idx, col_name_idx, col_price_idx, is_kopecks = detect_columns(rows)
    logger.info(f"[Price Parser Raw] col_article_idx={col_article_idx}, col_name_idx={col_name_idx}, col_price_idx={col_price_idx}, is_kopecks={is_kopecks}")

    raw_list = []
    for row in rows:
        if not any(row):
            continue

        art_val = row[col_article_idx] if col_article_idx < len(row) else None
        name_val = row[col_name_idx] if col_name_idx < len(row) else None
        price_val = row[col_price_idx] if col_price_idx < len(row) else None

        if not name_val:
            continue

        if name_val and str(name_val).strip().lower() in ["наименование", "наименование номенклатуры", "название", "номенклатура"]:
            continue

        # Parse price with robust float parser
        price = parse_robust_float(price_val)
        if is_kopecks:
            price /= 100.0

        price_str = str(price_val).strip() if price_val is not None else ""
        article_str = str(art_val).strip() if art_val is not None else ""

        # Validation rule 1: Verify article is not pure non-numeric category text (e.g. "Промышленная")
        if article_str and not re.search(r'\d', article_str):
            logger.error(f"[Price Validation Error] Article '{article_str}' is pure text without digits. Clearing article.")
            article_str = ""

        # Validation rule 2: Check if price equals article code (price == article)
        if article_str and (price_str == article_str or clean_key(article_str) == clean_key(price_str) or (price > 0 and str(int(price)) == article_str)):
            logger.error(f"[Price Validation Error] Parsed price ({price}) equals article code ({article_str}). Setting price to 0.0.")
            price = 0.0

        raw_list.append({
            "Артикул": article_str,
            "Наименование": str(name_val).strip(),
            "Тариф с НДС, руб": price,
            "price": price
        })

    return raw_list
