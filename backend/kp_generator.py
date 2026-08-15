from typing import List, Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import re
import logging
from backend.price_parser import clean_key
from smart_matcher import get_device_category

logger = logging.getLogger("kp_generator")

def strict_poles_current_match(poles: str, current_a: str, price_map: Dict[str, float]) -> float:
    """
    Strictly matches the price list based on poles and current_a values.
    Returns the price if found, else 0.0.
    """
    if not poles or not current_a:
        return 0.0

    poles_norm = poles.lower().strip() # e.g. "3p"

    # Extract only the numeric digits from current_a (e.g. "125" from "125А")
    current_digits_match = re.search(r'\d+', current_a)
    if not current_digits_match:
        return 0.0
    current_norm = current_digits_match.group(0) # e.g. "125"

    # Iterate over price keys to find strict matches
    for key, price in price_map.items():
        key_lower = key.lower()
        # 1. Verify poles match (e.g. "3p" is in the price key/desc)
        if poles_norm not in key_lower:
            continue
        # 2. Verify current matches strictly as a whole integer (not matching "16" in "160")
        pattern = r'(?<!\d)' + re.escape(current_norm) + r'(?!\d)'
        if re.search(pattern, key_lower):
            logger.info(f"[Pricing] Strict match found for poles={poles_norm}, current={current_norm} inside key: '{key}' (Price: {price})")
            return price

    return 0.0

def word_overlap_match(name: str, price_map: Dict[str, float]) -> tuple[float, bool]:
    """
    Looks for the best word-overlap match in the price map based on name keywords.
    Splits the item name into words (>= 3 chars) and counts intersections.
    """
    if not name or len(name) < 4:
        return 0.0, False

    words = [w.lower() for w in re.split(r'[^A-Za-z0-9А-Яа-я]', name) if len(w) >= 3]
    if not words:
        return 0.0, False

    best_key = None
    best_score = 0

    # We iterate over price keys
    for key in price_map:
        if not key or len(key) < 5:
            continue
        # Split price key/description into lowercase parts
        score = sum(1 for w in words if w in key.lower())
        if score > best_score:
            best_score = score
            best_key = key

    # We require at least 2 matching keywords for a solid match
    if best_key and best_score >= 2:
        return price_map[best_key], True

    return 0.0, False

def generate_preliminary_kp(boards: List[Dict[str, Any]], price_map: Dict[str, float], index_map: Dict[str, Any] = None) -> Dict[str, Any]:
    """
    Combines parsed board positions with pricing.
    - Strictly matches against active index `{poles}_{current_a}`.
    - If found: assigns correct article + price.
    - If not found: price = 0.0, article = "", name = "Авт. выкл. {poles} {current_a}А".
    - Ensures no schematic annotations like QF* are used as articles.
    """
    grand_total = 0.0
    kp_boards = []

    for board in boards:
        board_name = board["board_name"]
        board_items = []
        board_subtotal = 0.0

        for item in board["items"]:
            qty = item.get("qty", 1)
            unit = item.get("unit", "шт")
            poles = str(item.get("poles") or "").upper().strip()
            current_a = str(item.get("current_a") or "").strip()

            # Clean numerical current value
            current_digits_match = re.search(r'\d+', current_a)
            current_val = current_digits_match.group(0) if current_digits_match else ""

            price = 0.0
            price_found = False
            article = ""
            name = f"Авт. выкл. {poles} {current_val}А" if (poles and current_val) else str(item.get("name") or "")

            # Check if item has been already matched (e.g., by MATCHER in main.py)
            if item.get("price_found"):
                price = float(item.get("price") or 0.0)
                price_found = True
                art_candidate = str(item.get("article") or "").strip()
                if re.match(r'(?i)^QF\d+$', art_candidate):
                    article = ""
                else:
                    article = art_candidate
                name = str(item.get("name") or name)
                logger.info(f"[Match] pre-matched-fallback article={article} price={price}")

            elif poles and current_val:
                key = f"{poles}_{current_val}"
                if index_map and key in index_map and index_map[key]:
                    candidates = index_map[key]
                    best_cand = candidates[0]
                    best_score = -1

                    item_series = str(item.get("series") or "").strip().lower()
                    item_type = str(item.get("type") or "").strip().lower()
                    item_nominal = str(item.get("nominal") or "").strip().lower()
                    item_art = str(item.get("article") or item.get("mark") or "").strip().lower()
                    clean_item_art = clean_key(item_art)
                    item_cat = get_device_category(' '.join([str(item.get("name") or ""), item_series, item_type]))

                    for cand in candidates:
                        cand_name = str(cand.get("name") or cand.get("Наименование") or "").lower()
                        cand_art = str(cand.get("article") or cand.get("Артикул") or "").lower()
                        clean_cand_art = clean_key(cand_art)
                        cand_series = str(cand.get("series") or "").lower()
                        cand_cat = get_device_category(cand_name)

                        score = 0
                        # Check category conflict (e.g. breaker vs VFD)
                        if item_cat != 'unknown' and cand_cat != 'unknown' and item_cat != cand_cat:
                            score -= 100
                        else:
                            score += 10

                        # Exact article match bonus
                        if clean_item_art and clean_cand_art and clean_item_art == clean_cand_art:
                            score += 1000

                        # Nominal rating bonus (e.g. C16, D25, B10)
                        if item_nominal and (item_nominal in cand_name or item_nominal in cand_art):
                            score += 20

                        if item_series:
                            if item_series in cand_series or item_series in cand_name or item_series in cand_art:
                                score += 10
                            # Extract base series prefix (e.g., NM8N, NXM, NB2, NB8)
                            m_base = re.search(r'\b(nm8[ns]|nxm|nm1|nxb|nb[128]|nc[1278]|nvf[257]|nz7|nkb1|dz158|nr[8e]|nl1|nh4|nd2|nb1l)\b', item_series)
                            if m_base and m_base.group(1) in cand_name:
                                score += 5
                        if item_type and item_type in cand_name:
                            score += 2

                        if score > best_score:
                            best_score = score
                            best_cand = cand

                    matched_pos = best_cand
                    price = float(matched_pos.get("price") or 0.0)
                    art_candidate = str(matched_pos.get("article") or "").strip()
                    # Ensure no schematic label like QF* leaks into article
                    if re.match(r'(?i)^QF\d+$', art_candidate):
                        article = ""
                    else:
                        article = art_candidate
                    name = str(matched_pos.get("name") or matched_pos.get("Наименование") or name)
                    price_found = True
                    logger.info(f"[Match] key={key} article={article} price={price}")
                else:
                    logger.info(f"[Match] key={poles}_{current_val} article= price=0")
            else:
                logger.info(f"[Match] missing-poles-or-amperage price=0")

            # Compute total cost
            total_sum = price * qty
            board_subtotal += total_sum

            board_items.append({
                "article": article,
                "name": name,
                "qty": qty,
                "unit": unit,
                "price": round(price, 2),
                "total": round(total_sum, 2),
                "price_found": price_found
            })

        # Only add the board if it contains at least one parsed item
        if board_items:
            grand_total += board_subtotal
            kp_boards.append({
                "board_name": board_name,
                "items": board_items,
                "subtotal": round(board_subtotal, 2)
            })

    return {
        "boards": kp_boards,
        "grand_total": round(grand_total, 2)
    }

def export_kp_to_excel(kp_data: Dict[str, Any]) -> bytes:
    """
    Generates a highly polished and professional commercial offer (КП) spreadsheet.
    Includes styled headers, subtotals, borders, and total sums.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Коммерческое предложение"

    # Enable grid lines
    ws.views.sheetView[0].showGridLines = True

    # Styles
    font_title = Font(name="Calibri", size=16, bold=True, color="1E3A8A")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_regular = Font(name="Calibri", size=11)
    font_red = Font(name="Calibri", size=11, color="991B1B", italic=True)

    fill_header = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    fill_subtotal = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    fill_total = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")

    border_thin = Side(border_style="thin", color="D1D5DB")
    border_double = Side(border_style="double", color="1E3A8A")

    box_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    total_border = Border(top=border_thin, bottom=border_double)

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # Title Block
    ws.merge_cells("A1:G1")
    ws["A1"] = "ПРЕДВАРИТЕЛЬНОЕ КОММЕРЧЕСКОЕ ПРЕДЛОЖЕНИЕ"
    ws["A1"].font = font_title
    ws["A1"].alignment = align_center
    ws.row_dimensions[1].height = 40

    ws["A2"] = "Оборудование: Электрощитовое CHINT"
    ws["A2"].font = font_regular
    ws["A3"] = "Статус: Предварительный расчет"
    ws["A3"].font = font_regular

    # Table headers
    headers = ["№", "Артикул", "Наименование позиции", "Кол-во", "Ед. изм.", "Цена с НДС, руб.", "Сумма с НДС, руб."]
    start_row = 5

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = box_border

    ws.row_dimensions[start_row].height = 25

    current_row = start_row + 1
    global_idx = 1

    for board in kp_data.get("boards", []):
        # Group/Board header row
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
        board_cell = ws.cell(row=current_row, column=1, value=f"Раздел: {board['board_name']}")
        board_cell.font = font_bold
        board_cell.fill = fill_subtotal
        board_cell.alignment = align_left
        board_cell.border = Border(bottom=border_thin)
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # Items
        for item in board.get("items", []):
            ws.cell(row=current_row, column=1, value=global_idx).alignment = align_center
            ws.cell(row=current_row, column=2, value=item["article"]).alignment = align_left
            ws.cell(row=current_row, column=3, value=item["name"]).alignment = align_left

            qty_cell = ws.cell(row=current_row, column=4, value=item["qty"])
            qty_cell.alignment = align_right
            qty_cell.number_format = "#,##0"

            ws.cell(row=current_row, column=5, value=item["unit"]).alignment = align_center

            price_cell = ws.cell(row=current_row, column=6, value=item["price"])
            price_cell.alignment = align_right
            price_cell.number_format = "#,##0.00"

            total_cell = ws.cell(row=current_row, column=7, value=item["total"])
            total_cell.alignment = align_right
            total_cell.number_format = "#,##0.00"

            # Apply thin border to row cells and color unpriced elements
            for c in range(1, 8):
                cell = ws.cell(row=current_row, column=c)
                cell.border = box_border
                if not item["price_found"]:
                    cell.font = font_red
                else:
                    cell.font = font_regular

            ws.row_dimensions[current_row].height = 20
            current_row += 1
            global_idx += 1

        # Board Subtotal row
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        sub_cell_lbl = ws.cell(row=current_row, column=1, value=f"Итого по разделу '{board['board_name']}':")
        sub_cell_lbl.font = font_bold
        sub_cell_lbl.alignment = align_right
        sub_cell_lbl.fill = fill_subtotal

        sub_cell_val = ws.cell(row=current_row, column=7, value=board["subtotal"])
        sub_cell_val.font = font_bold
        sub_cell_val.alignment = align_right
        sub_cell_val.number_format = "#,##0.00"
        sub_cell_val.fill = fill_subtotal

        # borders for subtotal
        for c in range(1, 8):
            ws.cell(row=current_row, column=c).border = box_border

        ws.row_dimensions[current_row].height = 22
        current_row += 1

    # Grand Total row
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    total_cell_lbl = ws.cell(row=current_row, column=1, value="ВСЕГО К ОПЛАТЕ:")
    total_cell_lbl.font = font_bold
    total_cell_lbl.alignment = align_right
    total_cell_lbl.fill = fill_total

    total_cell_val = ws.cell(row=current_row, column=7, value=kp_data.get("grand_total", 0.0))
    total_cell_val.font = font_bold
    total_cell_val.alignment = align_right
    total_cell_val.number_format = "#,##0.00"
    total_cell_val.fill = fill_total

    for c in range(1, 8):
        ws.cell(row=current_row, column=c).border = total_border

    ws.row_dimensions[current_row].height = 26

    # Auto-adjust columns widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            # Skip title row and subtotal cells which are merged
            if cell.row in [1, 2, 3] or (cell.row > start_row and cell.column == 1 and "Раздел:" in str(cell.value)):
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    # Limit maximum description column width for readability
    ws.column_dimensions["C"].width = 45

    # Save Workbook to stream
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
