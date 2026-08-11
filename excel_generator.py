from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

class ExcelGenerator:
    def __init__(self, matched_items: List[Dict], total_cost: float):
        self.matched_items = matched_items
        self.total_cost = total_cost

    def generate(self) -> bytes:
        """Генерирует Excel-файл с коммерческим предложением"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Коммерческое предложение"

        # Заголовок
        ws['A1'] = "ПРЕДВАРИТЕЛЬНОЕ КОММЕРЧЕСКОЕ ПРЕДЛОЖЕНИЕ"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:G1')

        ws['A2'] = "Оборудование: Электрощитовое CHINT"
        ws['A3'] = "Статус: Предварительный расчет"

        # Шапка таблицы
        headers = ['№', 'Артикул', 'Наименование позиции', 'Кол-во', 'Ед. изм.', 'Цена с НДС, руб.', 'Сумма с НДС, руб.']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

        # Данные
        for idx, item in enumerate(self.matched_items, 1):
            row = idx + 5
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=item.get('article', ''))
            ws.cell(row=row, column=3, value=item.get('matched_name', item.get('name', '')))
            ws.cell(row=row, column=4, value=1)
            ws.cell(row=row, column=5, value='шт')
            ws.cell(row=row, column=6, value=item.get('price', 0))
            ws.cell(row=row, column=7, value=item.get('price', 0))

        # Итого
        last_row = len(self.matched_items) + 6
        ws.cell(row=last_row, column=6, value="ИТОГО:")
        ws.cell(row=last_row, column=7, value=self.total_cost)
        ws.cell(row=last_row, column=7).font = Font(bold=True)

        # Автоширина колонок
        for col in range(1, 8):
            ws.column_dimensions[chr(64 + col)].width = 25

        # Сохраняем в bytes
        import io
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
